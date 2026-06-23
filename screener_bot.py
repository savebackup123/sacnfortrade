# -*- coding: utf-8 -*-
"""
80/20 SET Screener Bot with Telegram Notification
Automatically downloads stock list, runs screening criteria, exports Excel,
and sends summary + file to Telegram.
"""

import os
import sys
import time
from io import StringIO
from datetime import datetime
import pandas as pd
import requests
import pandas_ta_classic as ta
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load environment variables manually from .env if it exists (for local testing)
if os.path.exists(".env"):
    with open(".env", "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, val = line.split("=", 1)
                os.environ[key.strip()] = val.strip()

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

def format_chat_id(chat_id):
    chat_id = str(chat_id).strip()
    if chat_id.startswith("-") and not chat_id.startswith("-100"):
        digits_part = chat_id[1:]
        if digits_part.isdigit() and len(digits_part) >= 9:
            return f"-100{digits_part}"
    return chat_id

def send_telegram_message(token, chat_id, text):
    if not token or not chat_id:
        print("Telegram configuration missing. Skip sending message.")
        return None
    chat_id = format_chat_id(chat_id)
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML"
    }
    try:
        response = requests.post(url, json=payload, timeout=30)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        print(f"Error sending telegram message: {e}")
        return None

def send_telegram_file(token, chat_id, filepath, caption):
    if not token or not chat_id:
        print("Telegram configuration missing. Skip sending file.")
        return None
    chat_id = format_chat_id(chat_id)
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    try:
        with open(filepath, "rb") as file:
            files = {"document": file}
            data = {"chat_id": chat_id, "caption": caption}
            response = requests.post(url, data=data, files=files, timeout=60)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        print(f"Error sending telegram file: {e}")
        return None

def main():
    today_str = datetime.now().strftime("%Y-%m-%d")
    print(f"=== Starting 80/20 SET Screener Bot - {today_str} ===")

    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("⚠️ Warning: TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID environment variables are not set.")

    # =========================
    # 1) GET SET TICKERS (exclude mai)
    # =========================
    print("Step 1: Fetching SET Tickers...")
    LIST_URL = "https://www.set.or.th/dat/eod/listedcompany/static/listedCompanies_th_TH.xls"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "th-TH,th;q=0.9,en;q=0.8",
    }
    
    tickers = []
    try:
        print("Attempting to download live tickers from SET website...")
        r = requests.get(LIST_URL, headers=headers, timeout=30)
        r.raise_for_status()
        df0 = pd.read_html(StringIO(r.text))[0]
        listed = df0.copy()
        listed.columns = listed.iloc[1]
        listed = listed.iloc[2:].reset_index(drop=True)
        listed = listed[["หลักทรัพย์", "ตลาด"]].copy()
        listed["หลักทรัพย์"] = listed["หลักทรัพย์"].astype(str).str.strip()
        listed["ตลาด"] = listed["ตลาด"].astype(str).str.strip()
        df_set = listed[listed["ตลาด"].str.upper().eq("SET")].copy()
        tickers = df_set["หลักทรัพย์"].dropna().astype(str).str.strip().unique().tolist()
        print(f"Successfully downloaded {len(tickers)} live tickers.")
    except Exception as e:
        print(f"⚠️ Warning: Failed to fetch tickers from SET website: {e}")
        if os.path.exists("set_tickers.json"):
            print("Using local backup set_tickers.json...")
            import json
            try:
                with open("set_tickers.json", "r", encoding="utf-8") as f:
                    tickers = json.load(f)
                print(f"Loaded {len(tickers)} tickers from set_tickers.json backup.")
            except Exception as json_err:
                print(f"❌ Error reading set_tickers.json: {json_err}")
        
    if not tickers:
        print("❌ Error: No tickers available to scan.")
        send_telegram_message(TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, "❌ <b>80/20 SET Screener Error:</b> ไม่สามารถโหลดรายชื่อหุ้นจากเว็บไซต์ SET หรือไฟล์ Backup ได้")
        sys.exit(1)

    tickers_bk = [t if t.endswith(".BK") else f"{t}.BK" for t in tickers]
    print(f"Total SET tickers to scan: {len(tickers_bk)}")

    # =========================
    # 2) RSI SCAN (ANY RSI>80 in last 75 trading days)
    # =========================
    print("Step 2: Performing RSI Scan (RSI > 80 in last 75 bars)...")
    PERIOD = "220d"      
    RSI_LEN = 14
    THRESH = 80
    LOOKBACK_BARS = 75

    def chunks(lst, n=120):
        for i in range(0, len(lst), n):
            yield lst[i:i+n]

    frames = []
    for idx, ch in enumerate(chunks(tickers_bk, 120), 1):
        print(f"Downloading batch {idx}...")
        price_df = yf.download(
            tickers=ch,
            period=PERIOD,
            interval="1d",
            group_by="ticker",
            auto_adjust=True,
            threads=True,
            progress=False
        )
        frames.append(price_df)

    raw = pd.concat(frames, axis=1)

    rows = []
    for t in tickers_bk:
        try:
            if t not in raw.columns.get_level_values(0):
                continue

            close = raw[t]["Close"].dropna()
            if len(close) < RSI_LEN + 10:
                continue

            rsi = ta.rsi(close, length=RSI_LEN).dropna()
            rsi_75 = rsi.tail(LOOKBACK_BARS)
            if len(rsi_75) == 0:
                continue

            rows.append({
                "Ticker": t,
                "Passed_ANY_RSI>80_in_75bars": (rsi_75 > THRESH).any(),
                "RSI_last": float(rsi_75.iloc[-1]),
                "RSI_max_75bars": float(rsi_75.max()),
                "Date_of_max_75bars": str(rsi_75.idxmax())[:10],
                "Bars_close": int(len(close))
            })
        except Exception:
            continue

    all_debug = pd.DataFrame(rows)
    passed_df = (
        all_debug[all_debug["Passed_ANY_RSI>80_in_75bars"]]
        .sort_values("RSI_max_75bars", ascending=False)
        .reset_index(drop=True)
    )

    print(f"Passed RSI filter: {len(passed_df)} tickers")

    # =========================
    # 3) SCREENING ต่อ (DY>5%, PE<12, EMA10>Close, EMA10>EMA50)
    # =========================
    print("Step 3: Screening fundamentals and EMAs...")
    MIN_DIV_YIELD = 0.05   
    MAX_PE = 12

    def calc_ema_flags_from_raw(raw, ticker_bk, ema_fast=10, ema_slow=50, lookback=220):
        if ticker_bk not in raw.columns.get_level_values(0):
            return None

        close = raw[ticker_bk]["Close"].dropna().tail(lookback)
        if len(close) < ema_slow + 5:
            return None

        ema10 = ta.ema(close, length=ema_fast).dropna()
        ema50 = ta.ema(close, length=ema_slow).dropna()
        if len(ema10) == 0 or len(ema50) == 0:
            return None

        close_last = float(close.iloc[-1])
        ema10_last = float(ema10.iloc[-1])
        ema50_last = float(ema50.iloc[-1])

        pullback = (ema10_last > close_last)   
        uptrend  = (ema10_last > ema50_last)   
        return close_last, ema10_last, ema50_last, pullback, uptrend

    def to_float(x):
        if x is None:
            return None
        v = pd.to_numeric(x, errors="coerce")
        if pd.isna(v):
            return None
        try:
            return float(v)
        except Exception:
            return None

    def normalize_dividend_yield(dy_raw):
        dy_num = to_float(dy_raw)
        if dy_num is None:
            return None, None
        if dy_num > 1:
            return dy_num / 100.0, "percent"
        else:
            return dy_num, "fraction"

    def get_fundamentals_yf(ticker_bk):
        try:
            info = yf.Ticker(ticker_bk).info
            dy_raw = info.get("dividendYield", None)
            pe_raw = info.get("trailingPE", None)

            dy_frac, dy_unit = normalize_dividend_yield(dy_raw)
            pe_num = to_float(pe_raw)

            return dy_raw, pe_raw, dy_frac, dy_unit, pe_num
        except Exception:
            return None, None, None, None, None

    screen_rows = []
    tickers_passed = passed_df["Ticker"].tolist()

    for i, t in enumerate(tickers_passed, 1):
        tech = calc_ema_flags_from_raw(raw, t, ema_fast=10, ema_slow=50, lookback=220)
        if tech is None:
            continue
        close_last, ema10_last, ema50_last, pullback, uptrend = tech

        dy_raw, pe_raw, dy_frac, dy_unit, pe = get_fundamentals_yf(t)

        missing_dy = (dy_frac is None)
        missing_pe = (pe is None)

        pass_dy = (dy_frac is not None) and (dy_frac > MIN_DIV_YIELD)
        pass_pe = (pe is not None) and (pe < MAX_PE)
        passed_all = pass_dy and pass_pe and pullback and uptrend

        dy_pct = None if dy_frac is None else dy_frac * 100.0

        screen_rows.append({
            "Ticker": t,
            "DY_raw": dy_raw,
            "PE_raw": pe_raw,
            "DividendYield_frac": dy_frac,
            "DividendYield_%": dy_pct,
            "DY_unit": dy_unit,
            "TrailingPE": pe,
            "Missing_DY": missing_dy,
            "Missing_PE": missing_pe,
            "Close_last": close_last,
            "EMA10_last": ema10_last,
            "EMA50_last": ema50_last,
            "EMA10>Close (Pullback)": pullback,
            "EMA10>EMA50 (Uptrend)": uptrend,
            "Pass_DY>5%": pass_dy,
            "Pass_PE<12": pass_pe,
            "Passed_ALL (DY>5%, PE<12, EMA)": passed_all
        })

        if i % 25 == 0:
            time.sleep(0.5)

    screen_df = pd.DataFrame(screen_rows)

    screened_df = (
        screen_df[screen_df["Passed_ALL (DY>5%, PE<12, EMA)"]]
        .copy()
        .sort_values(["DividendYield_%", "TrailingPE"], ascending=[False, True])
        .reset_index(drop=True)
    )

    print(f"Screened hits: {len(screened_df)}")

    # =========================
    # 4) README SHEET
    # =========================
    readme_df = pd.DataFrame([
        ["📌 OVERVIEW", "รายชื่อหุ้นในตลาด SET (ไม่รวม mai) ที่เคยมีค่า RSI(14) > 80 อย่างน้อย 1 วัน ภายใน 75 วันทำการล่าสุด"],
        ["", ""],
        ["🧠 CRITERIA (RSI)", "• ใช้ RSI(14)\n• ข้อมูลจาก Yahoo Finance\n• ใช้ 75 วันทำการจริง\n• เงื่อนไข: RSI > 80 อย่างน้อย 1 วัน"],
        ["", ""],
        ["📌 SCREENED_DY_PE_EMA", "แท็บ SCREENED_DY_PE_EMA คือผลการกรองต่อจาก RSI>80_ANY_75bars ด้วย:\n• Dividend Yield > 5%\n• P/E < 12\n• EMA(10) > Close (พักฐาน)\n• EMA(10) > EMA(50) (เทรนด์ขาขึ้น)\n\nหมายเหตุ: หุ้นที่ไม่มีข้อมูล DY หรือ P/E ใน yfinance จะถูกตัดออก (ถือว่าไม่ผ่าน)"],
        ["", ""],
        ["📊 COLUMN NOTES (SCREENING)", "• DY_raw / PE_raw = ค่าดิบจาก yfinance\n• DividendYield_frac = ค่า DY แบบ fraction (0.05 = 5%) ใช้ตัดสินเงื่อนไข\n• DividendYield_% = ค่า DY แบบเปอร์เซ็นต์เพื่ออ่านง่าย\n• DY_unit = บอกว่า yfinance ส่งมาเป็น fraction หรือ percent\n• Missing_DY / Missing_PE = ข้อมูลหาย/แปลงไม่ได้"],
    ], columns=["Item", "Description"])

    # =========================
    # 5) EXPORT + FORMAT EXCEL
    # =========================
    print("Step 4: Creating formatted Excel file...")
    out_xlsx = f"SET_RSI_over_80_any_in_last_75_trading_days_{today_str}.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        passed_df.to_excel(writer, index=False, sheet_name="RSI>80_ANY_75bars")
        all_debug.to_excel(writer, index=False, sheet_name="Debug_All")
        screened_df.to_excel(writer, index=False, sheet_name="SCREENED_DY_PE_EMA")
        screen_df.to_excel(writer, index=False, sheet_name="SCREEN_ALL_CANDIDATES")
        readme_df.to_excel(writer, index=False, sheet_name="README")

    wb = load_workbook(out_xlsx)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    def format_sheet(ws):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                cell.alignment = wrap
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    for name in ["RSI>80_ANY_75bars", "Debug_All", "SCREENED_DY_PE_EMA", "SCREEN_ALL_CANDIDATES", "README"]:
        format_sheet(wb[name])

    wb.save(out_xlsx)
    print(f"Excel saved as {out_xlsx}")

    # =========================
    # 6) SEND TO TELEGRAM
    # =========================
    print("Step 5: Sending results to Telegram...")
    tg_text = f"<b>📊 ผลการกรองหุ้น 80/20 SET Screener ({today_str})</b>\n"
    tg_text += "----------------------------------------------\n"
    
    if len(screened_df) > 0:
        tg_text += f"พบหุ้นผ่านเกณฑ์ทั้งหมด <b>{len(screened_df)}</b> ตัว:\n\n"
        for idx, row in screened_df.iterrows():
            ticker = row["Ticker"].replace(".BK", "")
            close = row["Close_last"]
            dy = row["DividendYield_%"]
            pe = row["TrailingPE"]
            tg_text += f"• <b>{ticker}</b>: ราคา <code>{close:.2f}</code> | ปันผล <code>{dy:.2f}%</code> | P/E <code>{pe:.2f}</code>\n"
        
        tg_text += "\n<i>*กรุณาตรวจสอบหน้ากราฟราคาเพื่อหาจังหวะซื้อตามรูปแบบ 80/20 อีกครั้ง</i>"
    else:
        tg_text += "ไม่พบหุ้นที่ผ่านเกณฑ์ครบทั้ง 4 มิติในวันนี้"

    tg_text += "\n\n📂 รายละเอียดเพิ่มเติมอยู่ในไฟล์ Excel แนบท้ายนี้"

    # Split chat IDs and send to each
    chat_ids = [c.strip() for c in TELEGRAM_CHAT_ID.split(",") if c.strip()] if TELEGRAM_CHAT_ID else []
    for cid in chat_ids:
        print(f"Sending notifications to Chat ID: {cid}")
        send_telegram_message(TELEGRAM_BOT_TOKEN, cid, tg_text)
        send_telegram_file(TELEGRAM_BOT_TOKEN, cid, out_xlsx, f"📊 ไฟล์ผลการสแกนหุ้นประจำวันที่ {today_str}")
    
    print("Telegram notifications sent successfully!")
    print("=== Finished ===")

if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as e:
        tb_str = traceback.format_exc()
        print(f"❌ Error occurred: {tb_str}")
        # Send error to Telegram to help user debug
        error_msg = f"❌ <b>80/20 SET Screener Run Error:</b>\n<pre>{tb_str[:3500]}</pre>"
        send_telegram_message(TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID, error_msg)
        sys.exit(1)
